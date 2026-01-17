import pathlib
from typing import Dict, List, Optional

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

DATA_PATH = pathlib.Path(__file__).parent / "data" / "SKU_WISE_AD_SPEND.xlsx"

TARGET_COLUMN_LETTERS = [
    "A",
    "B",
    "DV",
    "DX",
    "DZ",
    "EE",
    "EF",
    "EI",
    "EK",
    "EM",
    "EN",
    "EO",
]


@st.cache_data(show_spinner=False)
def load_column_mapping(path: pathlib.Path, letters: List[str]) -> Dict[str, str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in header_row]
    mapping: Dict[str, str] = {}
    for letter in letters:
        idx = column_index_from_string(letter) - 1
        if idx < 0 or idx >= len(headers):
            raise ValueError(f"Column letter {letter} is out of range.")
        column_name = headers[idx]
        if column_name is None:
            raise ValueError(f"Column letter {letter} has an empty header name.")
        mapping[letter] = str(column_name)
    return mapping


@st.cache_data(show_spinner=False)
def load_data(path: pathlib.Path, letters: List[str]) -> pd.DataFrame:
    mapping = load_column_mapping(path, letters)
    selected_columns = [mapping[letter] for letter in letters]
    df = pd.read_excel(path, sheet_name=0, usecols=selected_columns, engine="openpyxl")
    df.columns = [str(col).strip() for col in df.columns]
    for column in df.columns:
        if "date" in column.lower():
            df[column] = pd.to_datetime(df[column], errors="coerce")
    return df


def find_column(columns: List[str], keywords: List[str]) -> Optional[str]:
    for keyword in keywords:
        for column in columns:
            if keyword in column.lower():
                return column
    return None


def format_metric(value: Optional[float]) -> str:
    if value is None or pd.isna(value):
        return "N/A"
    if abs(value) >= 1_000_000:
        return f"{value / 1_000_000:.2f}M"
    if abs(value) >= 1_000:
        return f"{value / 1_000:.2f}K"
    return f"{value:,.2f}"


st.set_page_config(page_title="SKU WISE AD SPEND", page_icon="📊", layout="wide")

st.markdown(
    """
    <style>
        .kpi-card {background: #ffffff; border-radius: 12px; padding: 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.1);}
        .kpi-title {font-size: 0.85rem; color: #6c757d; margin-bottom: 4px;}
        .kpi-value {font-size: 1.5rem; font-weight: 600; color: #1f2a44;}
        .section-title {font-size: 1.1rem; font-weight: 600; color: #1f2a44; margin-bottom: 0.5rem;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("SKU WISE AD SPEND")
st.caption("Dashboard aligned to the Products Search Term 2025 experience, powered by SKU WISE AD SPEND data.")

if not DATA_PATH.exists():
    st.error("Dataset not found. Please ensure data/SKU_WISE_AD_SPEND.xlsx exists in the repository.")
    st.stop()

try:
    data = load_data(DATA_PATH, TARGET_COLUMN_LETTERS)
except Exception as exc:
    st.error(f"Failed to load dataset: {exc}")
    st.stop()

columns = list(data.columns)

numeric_columns = [col for col in columns if pd.api.types.is_numeric_dtype(data[col])]
date_columns = [col for col in columns if pd.api.types.is_datetime64_any_dtype(data[col])]

spend_col = find_column(columns, ["spend", "cost"])
sales_col = find_column(columns, ["sales", "revenue"])
orders_col = find_column(columns, ["order", "purchase"])
clicks_col = find_column(columns, ["click"])
impressions_col = find_column(columns, ["impression"])

st.sidebar.header("Filters")
filtered = data.copy()

if date_columns:
    date_column = date_columns[0]
    min_date = filtered[date_column].min()
    max_date = filtered[date_column].max()
    if pd.notna(min_date) and pd.notna(max_date):
        date_range = st.sidebar.date_input(
            "Date Range",
            value=(min_date.date(), max_date.date()),
            min_value=min_date.date(),
            max_value=max_date.date(),
        )
        if isinstance(date_range, tuple) and len(date_range) == 2:
            start_date, end_date = date_range
            filtered = filtered[(filtered[date_column] >= pd.Timestamp(start_date)) & (filtered[date_column] <= pd.Timestamp(end_date))]

categorical_columns = [
    col
    for col in columns
    if data[col].dtype == "object" and data[col].nunique(dropna=True) <= 30
]

for column in categorical_columns[:4]:
    options = sorted(filtered[column].dropna().unique().tolist())
    selection = st.sidebar.multiselect(column, options, default=options)
    if selection:
        filtered = filtered[filtered[column].isin(selection)]

if numeric_columns:
    for column in numeric_columns[:2]:
        series = filtered[column].dropna()
        if series.empty:
            continue
        min_val = float(series.min())
        max_val = float(series.max())
        if min_val != max_val:
            selected_range = st.sidebar.slider(
                f"{column} range",
                min_value=min_val,
                max_value=max_val,
                value=(min_val, max_val),
            )
            filtered = filtered[filtered[column].between(*selected_range)]

kpi_cols = st.columns(4)

spend_value = filtered[spend_col].sum() if spend_col else (filtered[numeric_columns[0]].sum() if numeric_columns else None)
sales_value = filtered[sales_col].sum() if sales_col else (filtered[numeric_columns[1]].sum() if len(numeric_columns) > 1 else None)
orders_value = filtered[orders_col].sum() if orders_col else (filtered[numeric_columns[2]].sum() if len(numeric_columns) > 2 else None)

roas_value = None
if spend_value is not None and sales_value is not None:
    roas_value = sales_value / spend_value if spend_value else None

kpi_labels = [
    ("Total Spend", spend_value),
    ("Total Sales", sales_value),
    ("Total Orders", orders_value),
    ("ROAS", roas_value),
]

for col, (label, value) in zip(kpi_cols, kpi_labels):
    with col:
        st.markdown(
            f"""
            <div class="kpi-card">
                <div class="kpi-title">{label}</div>
                <div class="kpi-value">{format_metric(value)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

st.markdown("<div class='section-title'>Performance Overview</div>", unsafe_allow_html=True)

left_col, right_col = st.columns(2)

with left_col:
    if date_columns and spend_col:
        trend_data = filtered.groupby(date_columns[0])[spend_col].sum().reset_index()
        fig = px.line(trend_data, x=date_columns[0], y=spend_col, markers=True)
        fig.update_layout(title="Spend Trend", margin=dict(l=20, r=20, t=40, b=20))
        st.plotly_chart(fig, use_container_width=True)
    elif spend_col and categorical_columns:
        group_col = categorical_columns[0]
        grouped = filtered.groupby(group_col)[spend_col].sum().sort_values(ascending=False).head(10).reset_index()
        fig = px.bar(grouped, x=group_col, y=spend_col, title=f"Top {group_col} by Spend")
        fig.update_layout(margin=dict(l=20, r=20, t=40, b=20))
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Not enough data to render the spend trend chart.")

with right_col:
    if spend_col and sales_col:
        fig = px.scatter(
            filtered,
            x=spend_col,
            y=sales_col,
            color=categorical_columns[0] if categorical_columns else None,
            title="Spend vs Sales",
        )
        fig.update_layout(margin=dict(l=20, r=20, t=40, b=20))
        st.plotly_chart(fig, use_container_width=True)
    elif clicks_col and impressions_col:
        fig = px.scatter(filtered, x=impressions_col, y=clicks_col, title="Impressions vs Clicks")
        fig.update_layout(margin=dict(l=20, r=20, t=40, b=20))
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Not enough data to render the performance scatter plot.")

st.markdown("<div class='section-title'>Detailed Table</div>", unsafe_allow_html=True)

st.dataframe(
    filtered,
    use_container_width=True,
    height=420,
    hide_index=True,
)
